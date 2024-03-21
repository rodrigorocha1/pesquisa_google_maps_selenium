
import sys
import os
from typing import List, Dict, Union
from openpyxl import Workbook, load_workbook, worksheet

from src.armazem.operacoes_arquivo import OperacaoArquivo

sys.path.insert(0, os.path.abspath(os.curdir))


class ArmazemExcel(OperacaoArquivo):

    def __init__(self, nome_arquivo: str, nome_aba: str) -> None:
        self.__planilha = Workbook()
        self.__nome_aba = nome_aba
        super().__init__(nome_arquivo)

    def __criar_cabecalho(self, dados: List[Dict[str, Union[str, int]]], aba: worksheet):
        cabecalhos = list(dados[0].keys())
        aba.append(cabecalhos)

        return cabecalhos

    def salvar_dados(self, dados: List[Dict[str, str]]):
        aba = self.__planilha.active
        aba.title = self.__nome_aba

        cabecalhos = self.__criar_cabecalho(dados=dados, aba=aba)
        for linha in dados:

            valores = [linha[coluna] for coluna in cabecalhos]
            aba.append(valores)
        self.__planilha.save(self._caminho_arquivo)
        self.__planilha.close()

    def atualizar_dados(self, dados: List[Dict[str, str]]):
        workbook = load_workbook(self._caminho_arquivo)
        if self.__nome_aba not in workbook.sheetnames:
            planilha = workbook.create_sheet(self.__nome_aba)
            cabecalhos = self.__criar_cabecalho(dados=dados, aba=planilha)
            for linha in dados:
                valores = [linha[coluna] for coluna in cabecalhos]
                planilha.append(valores)
        else:
            planilha = workbook[self.__nome_aba]
            ultima_lina = planilha.max_row + 1
            for _, valor in enumerate(dados, start=ultima_lina):
                planilha.append(list(valor.values()))

            ultima_lina = planilha.max_row

        workbook.save(self._caminho_arquivo)
        workbook.close()

    def verificar_arquivo(self):
        return os.path.exists(self._caminho_arquivo)
